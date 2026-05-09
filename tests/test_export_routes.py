import os
import sys
import types
from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from docx import Document
from openpyxl import load_workbook

os.environ.setdefault('WORKING_DIR', str(Path(__file__).resolve().parents[1]))
os.environ.setdefault('JWT_SECRET_KEY', 'test-secret-key')

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

if 'mem0' not in sys.modules:
    fake_mem0 = types.ModuleType('mem0')

    class _FakeMemory:
        def __init__(self, *args, **kwargs):
            pass

    fake_mem0.Memory = _FakeMemory
    sys.modules['mem0'] = fake_mem0

if 'phoenix' not in sys.modules:
    fake_phoenix = types.ModuleType('phoenix')

    class _FakePhoenixApp:
        url = 'http://localhost:6006'

        def close(self):
            return None

    fake_phoenix.launch_app = lambda *args, **kwargs: _FakePhoenixApp()
    sys.modules['phoenix'] = fake_phoenix

import main
from kuro_backend import chat_history, compliance_db, finance_db, intelligence_db


def _auth_client(monkeypatch, username='Pantronux') -> TestClient:
    monkeypatch.setattr(main, 'validate_token', lambda token: {'username': username})
    return TestClient(main.app)


def _seed_chat(chat_id='route_export_1', username='Pantronux', persona='advisor'):
    chat_history.create_session(chat_id, username, persona, title='Route Export Session')
    chat_history.add_message('web', 'user', 'Route user', [], persona, None, username, chat_id)
    chat_history.add_message('web', 'assistant', 'Route answer', [], persona, None, username, chat_id)


def _seed_intelligence(username='Pantronux'):
    intelligence_db.save_briefing(
        '2026-05-11',
        'Route intelligence summary',
        {'status_pagi': 'clear', 'wawasan_finansial': 'risk-on'},
        ['signal-1'],
        [{'symbol': 'TLKM.JK', 'stance': 'hold'}],
        username=username,
    )


def _seed_compliance():
    compliance_db.add_evidence('route_evidence.pdf', '/tmp/route_evidence.pdf', 'policy', 'iso27001', 'A.5')
    rows = compliance_db.get_evidence_matrix('iso27001')
    compliance_db.update_evidence_status(rows[0]['id'], 'compliant', 'good', 'keep')
    compliance_db.add_audit_trail('route_review', 'reviewed route evidence', 'iso27001')


def _seed_market(username='Pantronux'):
    finance_db.upsert_watched_symbol('BBRI.JK', 'BBRI', username=username)
    finance_db.apply_watched_price('BBRI.JK', 5000, username=username)
    finance_db.upsert_prediction_watch('usd-idr', 'USD/IDR Stress', 0.48, 'flat', username=username)
    finance_db.set_market_brief_and_note('Route market brief', 'Route sentinel note', username=username)


def test_post_export_md_returns_attachment(monkeypatch):
    client = _auth_client(monkeypatch)
    _seed_chat('route_export_md')
    response = client.post('/api/export', json={'target': 'chat_session', 'chat_id': 'route_export_md', 'format': 'md', 'message_ids': []}, cookies={main.COOKIE_NAME: 'Bearer dummy'})
    assert response.status_code == 200
    assert 'attachment;' in response.headers['content-disposition'].lower()
    assert '# Route Export Session' in response.text


def test_post_export_txt_returns_attachment(monkeypatch):
    client = _auth_client(monkeypatch)
    _seed_chat('route_export_txt')
    response = client.post('/api/export', json={'target': 'chat_session', 'chat_id': 'route_export_txt', 'format': 'txt', 'message_ids': []}, cookies={main.COOKIE_NAME: 'Bearer dummy'})
    assert response.status_code == 200
    assert 'Route user' in response.text


def test_post_export_json_returns_attachment(monkeypatch):
    client = _auth_client(monkeypatch)
    _seed_chat('route_export_json')
    response = client.post('/api/export', json={'target': 'chat_session', 'chat_id': 'route_export_json', 'format': 'json', 'message_ids': []}, cookies={main.COOKIE_NAME: 'Bearer dummy'})
    assert response.status_code == 200
    assert response.json()['title'] == 'Route Export Session'


def test_post_export_csv_returns_attachment(monkeypatch):
    client = _auth_client(monkeypatch)
    _seed_chat('route_export_csv')
    response = client.post('/api/export', json={'target': 'chat_session', 'chat_id': 'route_export_csv', 'format': 'csv', 'message_ids': []}, cookies={main.COOKIE_NAME: 'Bearer dummy'})
    assert response.status_code == 200
    assert 'text/csv' in response.headers['content-type']
    assert 'Route user' in response.text


def test_post_export_xlsx_returns_attachment(monkeypatch):
    client = _auth_client(monkeypatch)
    _seed_chat('route_export_xlsx')
    response = client.post('/api/export', json={'target': 'chat_session', 'chat_id': 'route_export_xlsx', 'format': 'xlsx', 'message_ids': []}, cookies={main.COOKIE_NAME: 'Bearer dummy'})
    assert response.status_code == 200
    workbook = load_workbook(filename=BytesIO(response.content))
    assert workbook.sheetnames == ['Metadata', 'Transcript']
    values = [workbook['Transcript'][f'F{row}'].value for row in range(2, 4)]
    assert 'Route user' in values
    assert 'Route answer' in values


def test_post_export_docx_returns_attachment(monkeypatch):
    client = _auth_client(monkeypatch)
    _seed_chat('route_export_docx')
    response = client.post('/api/export', json={'target': 'chat_session', 'chat_id': 'route_export_docx', 'format': 'docx', 'message_ids': []}, cookies={main.COOKIE_NAME: 'Bearer dummy'})
    assert response.status_code == 200
    document = Document(BytesIO(response.content))
    text = '\n'.join(paragraph.text for paragraph in document.paragraphs)
    assert 'Route Export Session' in text
    assert 'Route user' in text


def test_post_export_selected_messages_xlsx_returns_attachment(monkeypatch):
    client = _auth_client(monkeypatch)
    _seed_chat('route_export_selected_xlsx')
    messages = chat_history.get_history(chat_id='route_export_selected_xlsx', username='Pantronux', limit=9999)
    assistant_id = next(msg['id'] for msg in messages if msg['role'] == 'assistant')
    response = client.post(
        '/api/export',
        json={
            'target': 'selected_messages',
            'chat_id': 'route_export_selected_xlsx',
            'format': 'xlsx',
            'message_ids': [assistant_id],
        },
        cookies={main.COOKIE_NAME: 'Bearer dummy'},
    )
    assert response.status_code == 200
    workbook = load_workbook(filename=BytesIO(response.content))
    values = [workbook['Transcript'][f'F{row}'].value for row in range(2, 3)]
    assert 'Route answer' in values


def test_post_export_intelligence_json_returns_attachment(monkeypatch):
    client = _auth_client(monkeypatch)
    _seed_intelligence()
    response = client.post('/api/export', json={'target': 'intelligence_report', 'briefing_date': '2026-05-11', 'format': 'json'}, cookies={main.COOKIE_NAME: 'Bearer dummy'})
    assert response.status_code == 200
    data = response.json()
    assert data['export_type'] == 'intelligence_report'
    assert data['metadata']['briefing_date'] == '2026-05-11'


def test_post_export_market_json_returns_attachment(monkeypatch):
    client = _auth_client(monkeypatch)
    _seed_market()
    response = client.post('/api/export', json={'target': 'market_snapshot', 'format': 'json'}, cookies={main.COOKIE_NAME: 'Bearer dummy'})
    assert response.status_code == 200
    data = response.json()
    assert data['export_type'] == 'market_snapshot'
    assert data['metadata']['watched_symbol_count'] == '1'


def test_post_export_compliance_json_returns_attachment(monkeypatch):
    client = _auth_client(monkeypatch)
    _seed_compliance()
    response = client.post('/api/export', json={'target': 'compliance_report', 'standard': 'iso27001', 'format': 'json'}, cookies={main.COOKIE_NAME: 'Bearer dummy'})
    assert response.status_code == 200
    data = response.json()
    assert data['export_type'] == 'compliance_report'
    assert data['metadata']['standard'] == 'iso27001'


def test_post_export_compliance_denied_for_non_admin(monkeypatch):
    client = _auth_client(monkeypatch, username='Faikhira')
    _seed_compliance()
    response = client.post('/api/export', json={'target': 'compliance_report', 'standard': 'iso27001', 'format': 'json'}, cookies={main.COOKIE_NAME: 'Bearer dummy'})
    assert response.status_code == 403


def test_post_export_pdf_returns_accepted(monkeypatch):
    client = _auth_client(monkeypatch)
    _seed_chat('route_export_pdf')
    response = client.post('/api/export', json={'target': 'chat_session', 'chat_id': 'route_export_pdf', 'format': 'pdf', 'message_ids': []}, cookies={main.COOKIE_NAME: 'Bearer dummy'})
    assert response.status_code == 202
    assert response.json()['job_id'] > 0


def test_get_export_status_returns_404_for_unknown_job(monkeypatch):
    client = _auth_client(monkeypatch)
    response = client.get('/api/export/999999', cookies={main.COOKIE_NAME: 'Bearer dummy'})
    assert response.status_code == 404


def test_get_export_status_returns_403_for_other_user(monkeypatch):
    job_id = intelligence_db.create_export_job('OtherUser', 'chat_session', 'pdf', 'chat_xyz', [])
    client = _auth_client(monkeypatch)
    response = client.get(f'/api/export/{job_id}', cookies={main.COOKIE_NAME: 'Bearer dummy'})
    assert response.status_code == 403


def test_completed_pdf_download_returns_file(monkeypatch, tmp_path):
    client = _auth_client(monkeypatch)
    _seed_chat('route_export_download')
    monkeypatch.setattr(main.export_manager, 'EXPORT_ROOT', tmp_path)
    response = client.post('/api/export', json={'target': 'chat_session', 'chat_id': 'route_export_download', 'format': 'pdf', 'message_ids': []}, cookies={main.COOKIE_NAME: 'Bearer dummy'})
    job_id = response.json()['job_id']

    status_response = client.get(f'/api/export/{job_id}', cookies={main.COOKIE_NAME: 'Bearer dummy'})
    assert status_response.status_code == 200
    assert status_response.json()['status'] == 'completed'

    download_response = client.get(f'/api/export/{job_id}/download', cookies={main.COOKIE_NAME: 'Bearer dummy'})
    assert download_response.status_code == 200
    assert download_response.content.startswith(b'%PDF')


def test_failed_pdf_job_exposes_failed_status(monkeypatch):
    client = _auth_client(monkeypatch)
    _seed_chat('route_export_failed')

    def _explode(job_id):
        intelligence_db.mark_export_job_running(job_id)
        intelligence_db.mark_export_job_failed(job_id, 'boom')

    monkeypatch.setattr(main.export_manager, 'process_export_job', _explode)
    response = client.post('/api/export', json={'target': 'chat_session', 'chat_id': 'route_export_failed', 'format': 'pdf', 'message_ids': []}, cookies={main.COOKIE_NAME: 'Bearer dummy'})
    job_id = response.json()['job_id']
    status_response = client.get(f'/api/export/{job_id}', cookies={main.COOKIE_NAME: 'Bearer dummy'})
    assert status_response.status_code == 200
    assert status_response.json()['status'] == 'failed'
    assert status_response.json()['error_message'] == 'boom'


def test_legacy_export_md_still_works(monkeypatch):
    client = _auth_client(monkeypatch)
    _seed_chat('route_export_legacy_md')
    response = client.get('/api/chats/route_export_legacy_md/export?format=md', cookies={main.COOKIE_NAME: 'Bearer dummy'})
    assert response.status_code == 200
    assert response.headers['content-disposition'].endswith('.md"')


def test_legacy_export_txt_still_works(monkeypatch):
    client = _auth_client(monkeypatch)
    _seed_chat('route_export_legacy_txt')
    response = client.get('/api/chats/route_export_legacy_txt/export?format=txt', cookies={main.COOKIE_NAME: 'Bearer dummy'})
    assert response.status_code == 200
    assert response.headers['content-disposition'].endswith('.txt"')
