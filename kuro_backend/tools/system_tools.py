"""
Kuro AI V5.5 - System Tools for LangGraph Tool Node
================================================================================
Tools available:
- generate_excel_report: Create Excel files from JSON data
- manage_files: List, read, write, delete files in /home/kuro/exports/
- generate_report_template: Generate audit/compliance report templates
"""
import logging
import os
import json
from datetime import datetime
from typing import Dict, Any, Optional

logger = logging.getLogger(__name__)
logger.propagate = False  # Prevent double-reporting to root logger

# Tool descriptions for LLM routing
TOOL_DESCRIPTIONS = {
    "generate_excel_report": "Create an Excel spreadsheet from JSON data. Args: data (dict), filename (str), sheet_name (str)",
    "manage_files": "Manage files in /home/kuro/exports/. Actions: list, read, write, delete, info. Args: action (str), filename (str), content (str)",
    "generate_report_template": "Generate audit/compliance report templates. Args: template_type (str), filename (str), data (dict), format (str)",
}

EXPORTS_DIR = "/home/kuro/exports"
os.makedirs(EXPORTS_DIR, exist_ok=True)


def generate_excel_report(data: Optional[Dict] = None, filename: str = "report.xlsx", sheet_name: str = "Sheet1") -> Dict[str, Any]:
    """Generate an Excel report from JSON data."""
    try:
        import openpyxl
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Write headers
        if isinstance(data, dict):
            headers = list(data.keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Write data row
            for col, header in enumerate(headers, 1):
                ws.cell(row=2, column=col, value=str(data.get(header, "")))
        
        filepath = os.path.join(EXPORTS_DIR, filename if filename.endswith('.xlsx') else f"{filename}.xlsx")
        wb.save(filepath)
        
        return {"status": "success", "filepath": filepath, "message": f"Excel report saved to {filepath}"}
    except Exception as e:
        logger.error(f"Excel report generation failed: {e}")
        return {"status": "error", "message": str(e)}


def manage_files(action: str, filename: str = None, content: str = None) -> Dict[str, Any]:
    """Manage files in the exports directory."""
    try:
        if action == "list":
            files = os.listdir(EXPORTS_DIR) if os.path.exists(EXPORTS_DIR) else []
            return {"status": "success", "files": files, "count": len(files)}
        
        if not filename:
            return {"status": "error", "message": "Filename required for this action"}
        
        filepath = os.path.join(EXPORTS_DIR, filename)
        
        if action == "read":
            if not os.path.exists(filepath):
                return {"status": "error", "message": f"File not found: {filename}"}
            with open(filepath, 'r') as f:
                return {"status": "success", "content": f.read()}
        
        elif action == "write":
            if content is None:
                return {"status": "error", "message": "Content required for write action"}
            with open(filepath, 'w') as f:
                f.write(content)
            return {"status": "success", "message": f"File saved: {filepath}"}
        
        elif action == "delete":
            if os.path.exists(filepath):
                os.remove(filepath)
                return {"status": "success", "message": f"File deleted: {filename}"}
            return {"status": "error", "message": f"File not found: {filename}"}
        
        elif action == "info":
            if not os.path.exists(filepath):
                return {"status": "error", "message": f"File not found: {filename}"}
            stat = os.stat(filepath)
            return {
                "status": "success",
                "size": stat.st_size,
                "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            }
        
        return {"status": "error", "message": f"Unknown action: {action}"}
    except Exception as e:
        logger.error(f"File management failed: {e}")
        return {"status": "error", "message": str(e)}


def generate_report_template(template_type: str, filename: str, data: Optional[Dict] = None, format: str = "md") -> Dict[str, Any]:
    """Generate a report template for audit/compliance."""
    try:
        templates = {
            "audit_findings": """# Audit Findings Report

## Executive Summary
- **Date**: {date}
- **Auditor**: {auditor}
- **Scope**: {scope}

## Findings
| # | Finding | Severity | Status |
|---|---------|----------|--------|
| 1 |         |          |        |

## Recommendations
1. 
2. 
3. 
""",
            "compliance_report": """# Compliance Report

## Standard: {standard}
## Date: {date}

## Compliance Status
| Control | Status | Evidence |
|---------|--------|----------|
|         |        |          |

## Gaps Identified
1. 
2. 
""",
            "risk_assessment": """# Risk Assessment Report

## Date: {date}
## Assessor: {assessor}

## Risk Register
| Risk ID | Description | Likelihood | Impact | Risk Level |
|---------|-------------|------------|--------|------------|
|         |             |            |        |            |

## Mitigation Plan
1. 
2. 
""",
        }
        
        template = templates.get(template_type, "# Report Template\n\n## Content\n{content}")
        
        # Fill in common fields
        filled = template.format(
            date=datetime.now().strftime("%Y-%m-%d"),
            auditor=data.get("auditor", "") if data else "",
            scope=data.get("scope", "") if data else "",
            standard=data.get("standard", "") if data else "",
            assessor=data.get("assessor", "") if data else "",
            content=data.get("content", "") if data else "",
        )
        
        ext = format if format.startswith('.') else f".{format}"
        filepath = os.path.join(EXPORTS_DIR, filename if filename.endswith(ext) else f"{filename}{ext}")
        
        with open(filepath, 'w') as f:
            f.write(filled)
        
        return {"status": "success", "filepath": filepath, "message": f"Report template saved to {filepath}"}
    except Exception as e:
        logger.error(f"Report template generation failed: {e}")
        return {"status": "error", "message": str(e)}
